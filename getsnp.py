#! /usr/bin/env python3

# Anne-Sophie Denommé-Pichon
# 2020-01-04
# Script to get SNPs genotype by pedigree from LabKey export in xlsx

import openpyxl
import os.path
import sys

#{
#    'PED7699.1': {
#        '18069': {
#            '2010-01-02': {
#                'SNP10_rs9530': 'A/G',
#                'SNP11_rs20583': 'C/T'
#            }
#        }
#    }
#}

# Create an object (json like) to store the values extracted from the xlsx.
def load_genotypes_from_xlsx(input_path, ped_number):
    input_wb = openpyxl.load_workbook(input_path)
    ws = input_wb['data']

    values = {}
    snps = set()

    for row in ws.rows:
        if row[0].value.startswith(ped_number):
            snps.add(row[3].value)
            ped = values.setdefault(row[0].value, {})
            gad = ped.setdefault(str(row[14].value), {}) #FIXME : remove float type to GAD
            date = gad.setdefault(row[16].value, {})
            ## Genotype
            date[row[3].value] = row[9].value
    snps_names = sorted(snps)
    return values, snps_names
    
# Create an xlsx file to sumarize genotypes by PED number.
def save_genotypes_to_xlsx(values, snps_names, ped_number):
    output_wb = openpyxl.Workbook()
    ws = output_wb.active

    ## SNPs
    for index, snp in enumerate(snps_names):
        ws.cell(column=1, row=index + 4, value=snp)
    ## PED
    patient_column = 2
    for ped, gads in values.items(): #FIXME : sort by .2 .1 .3
        ws.cell(column=patient_column, row=1, value=ped)
        ## GAD
        for gad, dates in gads.items():
            ws.cell(column=patient_column, row=2, value=gad)
            ## Date
            for date, genotypes in dates.items():
                ws.cell(column=patient_column, row=3, value=date)
                ## Genotype
                for index, snp in enumerate(snps_names):
                    ws.cell(column=patient_column, row=index + 4, value=genotypes[snp])
                patient_column = patient_column + 1

    ws.column_dimensions['A'].width = 15.7
    for column in range(ord('B'), ord('Z')):
        ws.column_dimensions[chr(column)].width = 10
    
    output_wb.save(os.path.join(r"C:\Users\asden\Desktop", ped_number + ".xlsx"))
    
if __name__ == '__main__':
    if len(sys.argv) != 3:
        print(f'Usage: {sys.argv[0].split(os.sep)[-1]} PED_number input_file.xlsx')
        sys.exit(1)
    
    values, snps_names = load_genotypes_from_xlsx(sys.argv[2], sys.argv[1])
    save_genotypes_to_xlsx(values, snps_names, sys.argv[1])
    
